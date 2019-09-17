#__author__:YR
#__info__:None


import requests
import openpyxl
from requests.exceptions import RequestException
from bs4 import BeautifulSoup

def get_one_page(url, headers):
    try:
        respones=requests.get(url, headers=headers)
        if respones.status_code == 200:
            respones.encoding = "utf-8"
            return respones.text
        return None
    except RequestException:
        return None

def parser_one_page(html):
    soup = BeautifulSoup(html, "lxml")
    new_content = soup.find('div', class_='list').find_all("li")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "中国政府网新闻"
    ws["A1"] = "标题"
    ws["B1"] = "链接"
    ws["C1"] = "时间"
    for i, item in enumerate(new_content):
        news_title = item.find('a').get_text()
        news_href = item.find("a").get("href")
        if news_href.startswith("/zhengce"):
            news_href = "http://www.gov.cn" + news_href
        news_data = item.find('span').get_text().strip()
        ws.cell(row=i + 2, column=1, value=news_title)
        ws.cell(row=i + 2, column=2, value=news_href)
        ws.cell(row=i + 2, column=3, value=news_data)
    wb.save("zw_news.xlsx")

def main():
    url = "http://www.gov.cn/zhengce/zuixin.htm"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.119 Safari/537.36"}
    html = get_one_page(url, headers)
    parser_one_page(html)

if __name__ == "__main__":
    main()